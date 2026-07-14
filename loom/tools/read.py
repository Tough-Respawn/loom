# loom/tools/read.py
"""Outils de lecture : read_file (texte + PDF/Excel/Word, routage par extension) et
read_image. READ-only. Chemin absolu (n'importe où) ou relatif au dossier de travail.

UN SEUL outil de lecture de contenu : c'est read_file qui route selon l'extension
(extraction pour .pdf/.xlsx/.docx, décodage texte sinon) — jamais le modèle. L'ancien
read_document séparé faisait porter le routage au modèle, qui se trompait (« un .md est
un document ») et perdait un tour sur une erreur évitable."""

from __future__ import annotations

import base64
from pathlib import Path

from loom.agent.inline_image import wrap_image
from loom.tools.base import ToolError, ToolSpec, _resolve_in_root
from loom.tools.trust import untrusted

# Images servies au modèle multimodal (mmproj). MIME par extension.
_IMAGE_MIME = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".webp": "image/webp",
    ".bmp": "image/bmp",
}


def _decode_text_enc(data: bytes) -> tuple[str, str] | tuple[None, None]:
    """Décode des octets en (texte, encodage) en gérant les encodages courants sous
    Windows. L'encodage renvoyé permet de RÉ-ÉCRIRE à l'identique (edit_file). Renvoie
    (None, None) si ça ressemble vraiment à du binaire (échec de tout décodage)."""
    if data[:3] == b"\xef\xbb\xbf":  # UTF-8 avec BOM
        return data[3:].decode("utf-8", errors="replace"), "utf-8-sig"
    if data[:2] == b"\xff\xfe":  # UTF-16 LE (BOM) — défaut PowerShell
        return data[2:].decode("utf-16-le", errors="replace"), "utf-16"
    if data[:2] == b"\xfe\xff":  # UTF-16 BE (BOM)
        return data[2:].decode("utf-16-be", errors="replace"), "utf-16"
    try:
        return data.decode("utf-8"), "utf-8"  # cas le plus courant
    except UnicodeDecodeError:
        pass
    # UTF-16 sans BOM : beaucoup d'octets nuls (1 octet sur 2 pour de l'ASCII).
    if data and data.count(b"\x00") > len(data) // 4:
        for enc in ("utf-16-le", "utf-16-be"):
            try:
                return data.decode(enc), enc
            except UnicodeDecodeError:
                continue
    try:
        return data.decode("cp1252"), "cp1252"  # Windows-1252 (legacy), strict
    except UnicodeDecodeError:
        return None, None


def _decode_text(data: bytes) -> str | None:
    """Texte seul (compat) ; None si binaire. Cf. `_decode_text_enc` pour l'encodage."""
    return _decode_text_enc(data)[0]


def make_read_file(workspace_dir: str, max_bytes: int) -> ToolSpec:
    """Outil read_file : lit tout fichier commun (taille plafonnée, fenêtrable).

    Routage par extension, PAS par le modèle : .pdf/.xlsx/.xlsm/.docx -> extraction
    texte (`_run_doc`), image -> erreur qui pointe read_image, sinon texte brut.
    Plus d'allowlist d'extensions (Loom est généraliste : .env/.log/Dockerfile/CSV…
    sont légitimes) ; le garde anti-binaire de `_decode_text` rejette ce qui n'est pas
    du texte. Symétrique de write_file, qui n'a jamais restreint l'extension."""
    root = Path(workspace_dir)

    def run(args: dict) -> str:
        rel = (args.get("path") or "").strip()
        if not rel:
            raise ToolError("argument 'path' manquant")
        path = _resolve_in_root(root, rel)
        if not path.exists():
            raise ToolError(f"fichier introuvable : {rel}")
        if path.is_dir():
            raise ToolError(f"'{rel}' est un répertoire, pas un fichier")
        ext = path.suffix.lower()
        if ext in _DOC_READERS:
            return _run_doc(path, rel, args, max_bytes)
        if ext in _IMAGE_MIME:
            raise ToolError(f"'{rel}' est une image : utilise read_image")
        data = path.read_bytes()
        text = _decode_text(data)
        if text is None:
            raise ToolError(f"fichier binaire non lisible : {rel}")
        total_chars = len(text)
        if total_chars == 0:
            return f"[fichier vide : {rel}]"

        # Fenêtre de lecture (1-based, optionnelle) : lire un GROS fichier par TRANCHES
        # plutôt que tout d'un coup -> on ne dépasse jamais le contexte.
        # ALIAS tolérés : les modèles entraînés sur d'autres API écrivent souvent
        # `offset`/`limit`/`n`/`end_line`. Avant, ces clés étaient IGNORÉES en silence
        # (lecture depuis la ligne 1 en prétendant lire ailleurs). On les mappe.
        def _first(*keys):
            for k in keys:
                v = args.get(k)
                if v not in (None, ""):
                    return v
            return None

        sl = _first("start_line", "offset", "start", "from_line", "line")
        try:
            start = int(sl or 1)
        except (TypeError, ValueError):
            raise ToolError("start_line doit être un entier (1-based)") from None
        if start < 1:
            raise ToolError("start_line doit être >= 1 (1-based)")
        lc = _first("line_count", "limit", "n", "count", "num_lines")
        el = _first("end_line", "to_line")
        try:
            count = None if lc in (None, "") else int(lc)
            # `end_line` (borne haute inclusive) -> nombre de lignes, si line_count absent.
            if count is None and el not in (None, ""):
                count = max(0, int(el) - start + 1)
        except (TypeError, ValueError):
            raise ToolError("line_count/end_line doit être un entier") from None
        if count is not None and count < 1:
            # count=0 rendait une plage vide « lignes 2–1 » (silencieux, absurde).
            raise ToolError(
                "line_count doit être >= 1 (ou omis pour lire jusqu'au bout)"
            )
        sc = args.get("start_char")
        try:
            start_char = None if sc in (None, "") else int(sc)
        except (TypeError, ValueError):
            raise ToolError("start_char doit être un entier (1-based)") from None

        lines = text.splitlines()
        total = len(lines)
        if total == 0:
            return f"[fichier vide : {rel}]"

        # REPLI PAR CARACTÈRES : le découpage par lignes est inutile sur un fichier
        # MONO-LIGNE (JSON/CSS/JS minifié = une ligne géante) — le cap `max_bytes` ne
        # coupe qu'ENTRE lignes, donc une ligne de 74k passerait entière et saturerait le
        # contexte. On bascule en lecture par CARACTÈRES si `start_char` est fourni, OU
        # automatiquement dès qu'une ligne de la fenêtre dépasse le budget. Ainsi UNE
        # lecture ne renvoie JAMAIS plus de `max_bytes` caractères.
        if start > total and start_char is None:
            raise ToolError(f"start_line={start} hors fichier : {rel} a {total} lignes")
        end_check = total if count is None else min(total, start - 1 + count)
        longest = max((len(lines[i]) for i in range(start - 1, end_check)), default=0)
        if start_char is not None or longest > max_bytes:
            if start_char is None:
                start_char = 1  # auto-bascule : on lit depuis le début, par caractères
            if start_char < 1:
                raise ToolError("start_char doit être >= 1 (1-based)")
            if start_char > total_chars:
                raise ToolError(
                    f"start_char={start_char} hors fichier : {rel} a {total_chars} caractères"
                )
            chunk = text[start_char - 1 : start_char - 1 + max_bytes]
            end_char = start_char - 1 + len(chunk)
            head = (
                f"[LECTURE PAR CARACTÈRES — {rel} est peu/pas découpé en lignes (ex. minifié) : "
                "je le lis par tranches de caractères, pas par lignes.]\n"
            )
            if end_char >= total_chars:
                footer = (
                    f"\n[FIN DU FICHIER — caractères {start_char}–{total_chars} sur "
                    f"{total_chars}.]"
                )
            else:
                footer = (
                    f"\n[affiché caractères {start_char}–{end_char} sur {total_chars} — le "
                    f"FICHIER continue. Lis la SUITE avec read_file(path, "
                    f"start_char={end_char + 1}).]"
                )
            return head + chunk + footer

        # MODE LIGNES : sélection bornée par line_count ET par le cap de caractères. On
        # coupe AVANT de dépasser `max_bytes` (jamais d'overshoot d'une ligne entière ; les
        # lignes plus grosses que le budget sont déjà parties en mode caractères ci-dessus).
        end_limit = total if count is None else min(total, start - 1 + count)
        selected: list[str] = []
        chars = 0
        i = start - 1
        while i < end_limit:
            if selected and chars + len(lines[i]) + 1 > max_bytes:
                break
            selected.append(lines[i])
            chars += len(lines[i]) + 1
            i += 1
        last = i  # index exclusif -> dernière ligne affichée = i (1-based)
        # NUMÉROS DE LIGNE (format `  12→contenu`), ABSOLUS : le modèle les référence pour
        # se repérer/naviguer. Pour ÉDITER : copie l'extrait EXACT (sans le préfixe
        # `N→`) dans edit_file(old_string).
        width = max(2, len(str(last)))
        numbered = "\n".join(
            f"{n:>{width}}→{line}" for n, line in enumerate(selected, start)
        )
        # Marqueur de fin EXPLICITE : sans lui, le modèle qui voit du code s'arrêter net
        # croit que SA LECTURE a été coupée et relit en boucle. Si tronqué, on indique la
        # COMMANDE pour lire la suite (mécanisme réel, pas un vœu pieux).
        if last >= total:
            footer = (
                f"\n[FIN DU FICHIER — lignes {start}–{total} sur {total}. Pour éditer : "
                "edit_file (copie l'extrait exact à changer). Si le code s'arrête net, "
                "c'est LE FICHIER qui est incomplet (pas ta lecture) : complète-le "
                "(append_file pour la fin) — ne relis pas en boucle.]"
            )
        else:
            footer = (
                f"\n[affiché lignes {start}–{last} sur {total} — le FICHIER est plus long. "
                f"Lis la suite avec read_file(path, start_line={last + 1}), ou cible une "
                "zone précise avec search_text puis read_file(start_line=…).]"
            )
        return numbered + footer

    return ToolSpec(
        name="read_file",
        description=(
            "Reads ANY common file and returns its content: text/code (with line "
            "numbers), and DOCUMENTS — PDF (.pdf), Excel (.xlsx/.xlsm), Word (.docx) — "
            "whose text is extracted automatically (invoice, spreadsheet, report…). "
            "Path relative to the working directory OR absolute (e.g. "
            "'C:/Users/moi/Desktop/notes.txt'). Only images go elsewhere: read_image. "
            "LARGE file: read it in CHUNKS with start_line (and optionally line_count) — "
            "the response footer tells you where to continue. MINIFIED file (JSON/CSS/JS "
            "on a single line): reading switches AUTOMATICALLY to characters; then "
            "continue with start_char (the footer gives you the value). To target a "
            "precise area, first run search_text then read_file(start_line=…)."
        ),
        parameters={
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": (
                        "File path: relative to the working directory or absolute "
                        "(e.g. 'C:/Users/moi/notes.txt')."
                    ),
                },
                "start_line": {
                    "type": "integer",
                    "description": (
                        "First line to read (1-based, default 1). To read the rest of a "
                        "large file already started."
                    ),
                },
                "line_count": {
                    "type": "integer",
                    "description": (
                        "Number of lines to read from start_line (default: to the end or "
                        "the size limit)."
                    ),
                },
                "start_char": {
                    "type": "integer",
                    "description": (
                        "Start position IN CHARACTERS (1-based) for MINIFIED/single-line "
                        "files where start_line splits nothing. Generally unnecessary: the "
                        "switch is automatic — reuse the value given in the footer "
                        "'read the rest with start_char=…'."
                    ),
                },
            },
            "required": ["path"],
        },
        run=run,
    )


def _read_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    return "\n\n".join((page.extract_text() or "") for page in reader.pages)


def _read_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    out: list[str] = []
    try:
        for ws in wb.worksheets:
            out.append(f"# Feuille : {ws.title}")
            for row in ws.iter_rows(values_only=True):
                cells = ["" if c is None else str(c) for c in row]
                if any(cells):
                    out.append("\t".join(cells))
    finally:
        wb.close()
    return "\n".join(out)


def _read_docx(path: Path) -> str:
    import docx

    return "\n".join(p.text for p in docx.Document(str(path)).paragraphs)


_DOC_READERS = {
    ".pdf": _read_pdf,
    ".xlsx": _read_xlsx,
    ".xlsm": _read_xlsx,
    ".docx": _read_docx,
}


def _run_doc(path: Path, rel: str, args: dict, max_bytes: int) -> str:
    """Branche DOCUMENT de read_file : extrait le TEXTE d'un PDF / Excel / Word.

    Fenêtrage par CARACTÈRES uniquement (start_char) : le texte extrait ne s'édite
    pas, les numéros de ligne n'y serviraient qu'à tromper edit_file. Contenu reçu
    de l'extérieur (facture, rapport) -> encadré par la frontière de confiance.
    Lazy-import des libs (pypdf/openpyxl/python-docx) : message clair si une lib
    manque, jamais de crash."""
    ext = path.suffix.lower()
    try:
        text = _DOC_READERS[ext](path).strip()
    except ImportError as exc:
        raise ToolError(f"bibliothèque manquante pour {ext} : {exc}") from exc
    except Exception as exc:  # noqa: BLE001 - document corrompu/protégé
        raise ToolError(f"lecture impossible de {rel} : {exc}") from exc
    if not text:
        return f"(document sans texte extractible : {rel})"
    total = len(text)
    sc = args.get("start_char")
    try:
        start = 1 if sc in (None, "") else int(sc)
    except (TypeError, ValueError):
        raise ToolError("start_char doit être un entier (1-based)") from None
    if start < 1:
        raise ToolError("start_char doit être >= 1 (1-based)")
    if start > total:
        raise ToolError(
            f"start_char={start} hors document : {rel} a {total} caractères"
        )
    chunk = text[start - 1 : start - 1 + max_bytes]
    end = start - 1 + len(chunk)
    if end >= total:
        footer = f"\n[FIN DU DOCUMENT — caractères {start}–{total} sur {total}.]"
    else:
        footer = (
            f"\n[affiché caractères {start}–{end} sur {total} — le DOCUMENT continue. "
            f"Lis la suite avec read_file(path, start_char={end + 1}).]"
        )
    return untrusted(chunk, f"document {rel}") + footer


def make_read_image(
    workspace_dir: str,
    max_bytes: int = 10 * 1024 * 1024,
    describer=None,
    active_is_vision: bool = True,
) -> ToolSpec:
    """Outil read_image : fait accéder le modèle à une image du disque.

    - Modèle VISION (mmproj / VLM natif, active_is_vision=True) : l'image ne peut pas
      transiter par un message `tool` (texte seul) -> on renvoie une sentinelle (cf.
      loom.inline_image) que la boucle convertit en message `user` multimodal (il la VOIT).
    - Modèle TEXTE-ONLY (active_is_vision=False) avec un `describer` (routage vers un VLM) :
      le modèle ne voit rien, alors read_image lui renvoie une DESCRIPTION texte produite par
      le VLM (approche « VLM comme outil » : il interroge l'image à la demande, avec une
      question ciblée). Sans describer configuré : erreur claire.
    """
    root = Path(workspace_dir)

    def run(args: dict) -> str:
        rel = (args.get("path") or "").strip()
        if not rel:
            raise ToolError("argument 'path' manquant")
        path = _resolve_in_root(root, rel)
        if not path.exists():
            raise ToolError(f"fichier introuvable : {rel}")
        if path.is_dir():
            raise ToolError(f"'{rel}' est un répertoire, pas une image")
        mime = _IMAGE_MIME.get(path.suffix.lower())
        if mime is None:
            raise ToolError(
                f"format image non géré ({path.suffix or 'aucune extension'}) : "
                "png/jpg/jpeg/gif/webp/bmp"
            )
        data = path.read_bytes()
        if len(data) > max_bytes:
            raise ToolError(
                f"image trop volumineuse ({len(data)} octets > {max_bytes})"
            )
        b64 = base64.b64encode(data).decode("ascii")
        data_uri = f"data:{mime};base64,{b64}"
        # Modèle qui ne voit pas : réponse FRANCHE (règle : read_image = le modèle en
        # cours, jamais un détour vers un autre modèle — loom.web passe describer=None).
        if not active_is_vision:
            if describer is None:
                raise ToolError(
                    "ce modèle N'A PAS la vision : il ne peut pas lire d'image. "
                    "Dis-le franchement à l'utilisateur et propose-lui de basculer sur "
                    "un modèle marqué VISION dans le sélecteur (infobulle au survol)."
                )
            desc = describer(data_uri, (args.get("question") or "").strip())
            return untrusted(desc, f"image {rel} (décrite par le modèle vision)")
        return wrap_image(data_uri, rel)

    return ToolSpec(
        name="read_image",
        description=(
            "Accesses an IMAGE ON DISK (png/jpg/jpeg/gif/webp/bmp) from its PATH. If you "
            "are a multimodal model, you SEE it; otherwise, a vision model DESCRIBES it to "
            "you in text (you can ask a targeted `question`: 'what is the exact text?', "
            "'describe the layout'). For a file present on disk whose path you know "
            "(relative to the working directory or absolute), including images attached to "
            "the chat that are flagged to you with their path. Use it to describe, read "
            "text, compare a rendering. Any NON-image file (text, PDF/Excel/Word) -> read_file."
        ),
        parameters={
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": (
                        "Image path: relative to the working directory or absolute "
                        "(e.g. 'C:/Users/moi/Desktop/capture.png')."
                    ),
                },
                "question": {
                    "type": "string",
                    "description": (
                        "Optional: what you want to know about the image (ignored if you "
                        "see it directly; used to target the description when a vision "
                        "model describes it for you)."
                    ),
                },
            },
            "required": ["path"],
        },
        run=run,
    )
