# loom/tools/calc.py
"""Outil calculate : arithmétique EXACTE + agrégats sur données tabulaires.

Pourquoi : un LLM calcule « de tête » par prédiction de tokens et se trompe (vécu
2026-07-10 : somme fausse dans une analyse financière — le modèle lisait un CSV puis
additionnait mentalement 40 lignes). Même logique que format_code : un outil DÉDIÉ,
sûr et nommé, plutôt qu'espérer un `python -c` correct via run_shell.

UN SEUL réflexe pour le modèle (décision 2026-07-10, plutôt qu'un second outil) :
- sans `file`  : expression arithmétique pure — (55000-600)*0.2 ;
- avec `file`  : les chaînes deviennent des RÉFÉRENCES DE COLONNES du CSV/XLSX et les
  agrégats s'activent — sum("Débit") - sum("Crédit"), avg("Montant"), count("Compte").
  Filtre d'égalité optionnel (`where`) : somme des débits du compte 512 uniquement.

Évaluateur AST STRICT : nombres, opérateurs, parenthèses, fonctions whitelistées.
Aucun accès aux noms/attributs/appels hors liste -> rien d'autre n'est exécutable,
d'où : pas de garde de permission.
"""

from __future__ import annotations

import ast
import csv
import math
import re
from pathlib import Path

from loom.tools.base import ToolError, ToolSpec, _resolve_in_root

# --- table (CSV / XLSX) ---------------------------------------------------


class _Col:
    """Référence de colonne (issue d'une chaîne dans l'expression, mode fichier)."""

    __slots__ = ("name", "values")

    def __init__(self, name: str, values: list):
        self.name = name
        self.values = values


def _to_number(cell) -> float | None:
    """Coercition TOLÉRANTE (exports réels : « 1 234,56 € », espaces insécables).
    None = non numérique (vide, texte) — les agrégats l'ignorent en le comptant."""
    if cell is None:
        return None
    if isinstance(cell, (int, float)) and not isinstance(cell, bool):
        return float(cell)
    s = str(cell).strip().replace(" ", "").replace("\xa0", "").replace(" ", "")
    for sym in ("€", "$", "£"):
        s = s.replace(sym, "")
    if not s:
        return None
    # Décimale française : virgule SANS point -> point ; « 1.234,56 » -> 1234.56.
    if "," in s:
        s = s.replace(".", "").replace(",", ".") if s.count(",") == 1 else s
    try:
        return float(s)
    except ValueError:
        return None


def _load_table(path: Path) -> tuple[list[str], list[list]]:
    """(en-têtes, lignes) depuis un CSV (délimiteur détecté : ; , tab) ou un XLSX
    (première feuille). Erreurs ACTIONNABLES, jamais d'exception opaque."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ToolError("lecture xlsx impossible : openpyxl manquant") from exc
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001 - fichier corrompu/verrouillé
            raise ToolError(f"xlsx illisible : {exc}") from exc
        try:
            rows = [list(r) for r in wb.active.iter_rows(values_only=True)]
        finally:
            wb.close()  # sinon le handle reste ouvert -> fichier verrouillé (Windows)
    elif suffix in (".csv", ".tsv", ".txt"):
        try:
            text = path.read_text(encoding="utf-8-sig", errors="replace")
        except OSError as exc:
            raise ToolError(f"fichier illisible : {exc}") from exc
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        except csv.Error:

            class dialect(csv.excel):  # repli : virgule
                pass

        rows = list(csv.reader(text.splitlines(), dialect))
    else:
        raise ToolError(f"format non supporté : {suffix} (CSV, TSV ou XLSX)")
    rows = [r for r in rows if any(c not in (None, "") for c in r)]
    if not rows:
        raise ToolError("fichier vide")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return headers, rows[1:]


def _column(headers: list[str], rows: list[list], name: str) -> _Col:
    wanted = name.strip().lower()
    for i, h in enumerate(headers):
        if h.lower() == wanted:
            return _Col(headers[i], [r[i] if i < len(r) else None for r in rows])
    raise ToolError(
        f"colonne inconnue : {name!r} (colonnes du fichier : {', '.join(headers)})"
    )


def _apply_where(headers, rows, where: dict) -> list[list]:
    col = str(where.get("column", "")).strip()
    if not col or "equals" not in where:
        raise ToolError('`where` attend {"column": "...", "equals": "..."}')
    ref = _column(headers, rows, col)
    idx = headers.index(ref.name)
    target = where["equals"]
    t_num = _to_number(target)
    t_str = str(target).strip().lower()
    out = []
    for r in rows:
        cell = r[idx] if idx < len(r) else None
        c_num = _to_number(cell)
        if (t_num is not None and c_num is not None and c_num == t_num) or (
            str(cell).strip().lower() == t_str
        ):
            out.append(r)
    return out


# --- normalisation des écritures « modèle » -------------------------------
# Un LLM écrit la puissance de dix façons (^ , exposants Unicode e²², ×, ÷, moins
# Unicode −). Python ne connaît que ** , * , / , - -> on TRADUIT ces conventions en
# syntaxe Python AVANT le parse, HORS des chaînes entre guillemets (un nom de colonne
# « a^b » en mode fichier ne doit pas être réécrit). Objectif : que toute écriture
# usuelle « juste marche », sans que le modèle ait à connaître notre grammaire.

# Exposants Unicode -> chiffre/signe ASCII (⁰¹²³… ⁺ ⁻). Un run devient **(…).
_SUPERSCRIPT = {
    "⁰": "0", "¹": "1", "²": "2", "³": "3", "⁴": "4", "⁵": "5", "⁶": "6",
    "⁷": "7", "⁸": "8", "⁹": "9", "⁺": "+", "⁻": "-", "⁽": "(", "⁾": ")",
}  # fmt: skip

# Opérateurs/symboles Unicode -> équivalent ASCII Python.
_UNICODE_OPS = {
    "×": "*", "·": "*", "⋅": "*", "∗": "*", "∙": "*",
    "÷": "/", "∕": "/", "⁄": "/",
    "−": "-", "‑": "-", "–": "-", "—": "-",  # moins/tirets Unicode
    "π": "pi", "τ": "tau",  # symboles maths courants (√ traité à part : wrapping)
    "^": "**",
}  # fmt: skip


def _normalize_expr(expr: str) -> str:
    """Traduit les conventions d'écriture d'un modèle en syntaxe Python, hors chaînes
    entre guillemets. `^` et les exposants Unicode -> `**` (priorité correcte, car un
    remplacement textuel adopte la priorité forte de `**`, contrairement à un remappage
    du nœud XOR). `×·÷−` -> `*/-`, `π√` -> `pi sqrt`."""
    out: list[str] = []
    quote: str | None = None
    i, n = 0, len(expr)
    while i < n:
        ch = expr[i]
        if quote is not None:
            out.append(ch)
            if ch == quote:
                quote = None
            i += 1
            continue
        if ch in ("'", '"'):
            quote = ch
            out.append(ch)
            i += 1
            continue
        if ch in _SUPERSCRIPT:
            # Un run d'exposants Unicode (ex. « ²² » de e²²) -> **(22).
            j = i
            run = []
            while j < n and expr[j] in _SUPERSCRIPT:
                run.append(_SUPERSCRIPT[expr[j]])
                j += 1
            out.append("**(" + "".join(run) + ")")
            i = j
            continue
        out.append(_UNICODE_OPS.get(ch, ch))
        i += 1
    return _rewrite_math_shorthand("".join(out))


# Notation « raccourcie » d'un modèle, appliquée HORS chaînes (déjà séparées ci-dessus).
_NUM = r"\d*\.?\d+"


def _rewrite_math_shorthand(s: str) -> str:
    """√ sans parenthèses, multiplication implicite, `%` postfixe (pourcentage). Chaque
    passe est bornée pour NE PAS casser un identifiant (log10) ni la notation
    scientifique (1e3) : cf. les lookarounds."""
    # √ : √16 -> sqrt(16) ; √(x) -> sqrt(x) ; √ isolé -> sqrt
    s = re.sub(r"√\s*\(", "sqrt(", s)
    s = re.sub(rf"√\s*({_NUM})", r"sqrt(\1)", s)
    s = s.replace("√", "sqrt")
    # `%` postfixe = pourcentage : 20% -> (20*0.01). PAS quand un opérande suit (modulo :
    # 10%3 reste 10%3) — d'où le lookahead négatif sur un chiffre/point/paren.
    s = re.sub(rf"({_NUM})\s*%(?!\s*[\d.(])", r"(\1*0.01)", s)
    # Multiplication implicite : 2pi -> 2*pi, 3(4+5) -> 3*(4+5), )(  -> )*( .
    # Lookbehind : le nombre ne doit pas être la fin d'un identifiant (log10). Lookahead
    # (?![eE][+-]?\d) : ne pas couper la notation scientifique 1e3 / 1.5e-3.
    s = re.sub(
        rf"(?<![A-Za-z0-9_.])({_NUM})(?![eE][+-]?\d)\s*(?=[A-Za-z_(])", r"\1*", s
    )
    s = re.sub(r"(\))\s*(?=[A-Za-z0-9_(])", r"\1*", s)
    return s


# --- évaluateur AST -------------------------------------------------------

_FUNCS = {
    # arithmétique de base
    "abs": abs,
    "fabs": math.fabs,
    "round": round,
    "min": min,
    "max": max,
    "pow": pow,
    "sign": lambda x: (x > 0) - (x < 0),
    "copysign": math.copysign,
    "fmod": math.fmod,
    # racines / arrondis
    "sqrt": math.sqrt,
    "cbrt": math.cbrt,
    "floor": math.floor,
    "ceil": math.ceil,
    "trunc": math.trunc,
    # logs / exponentielles
    "log": math.log,  # log(x) naturel, ou log(x, base)
    "ln": math.log,
    "log10": math.log10,
    "log2": math.log2,
    "log1p": math.log1p,
    "exp": math.exp,
    "expm1": math.expm1,
    # combinatoire / entiers
    "factorial": math.factorial,
    "gcd": math.gcd,
    "lcm": math.lcm,
    "comb": math.comb,
    "perm": math.perm,
    "hypot": math.hypot,
    # trigonométrie
    "sin": math.sin,
    "cos": math.cos,
    "tan": math.tan,
    "asin": math.asin,
    "acos": math.acos,
    "atan": math.atan,
    "atan2": math.atan2,
    "sinh": math.sinh,
    "cosh": math.cosh,
    "tanh": math.tanh,
    "asinh": math.asinh,
    "acosh": math.acosh,
    "atanh": math.atanh,
    "degrees": math.degrees,
    "radians": math.radians,
}

# Agrégats : n'acceptent QUE des colonnes (mode fichier). `values` filtrés du non-num.
_AGGS = {
    "sum": lambda v: sum(v),
    "avg": lambda v: sum(v) / len(v) if v else 0.0,
    "mean": lambda v: sum(v) / len(v) if v else 0.0,
    "count": len,
    "min": min,
    "max": max,
}

_CONSTS = {"pi": math.pi, "e": math.e, "tau": math.tau, "inf": math.inf}


def _safe_pow(a, b):
    """Puissance BORNÉE : `9^9^9` (exposant = 387 M) produit un entier de centaines de
    millions de chiffres qui GÈLE le process (calc n'a aucun timeout). On refuse un
    exposant démesuré avec une erreur claire, au lieu de partir en calcul infini."""
    if isinstance(b, (int, float)) and abs(b) > 1e4 and abs(a) > 1:
        raise ToolError(
            "exposant trop grand (calcul refusé pour éviter un gel) : "
            f"|{b}| dépasse 10000"
        )
    return a**b


_BINOPS = {
    ast.Add: lambda a, b: a + b,
    ast.Sub: lambda a, b: a - b,
    ast.Mult: lambda a, b: a * b,
    ast.Div: lambda a, b: a / b,
    ast.FloorDiv: lambda a, b: a // b,
    ast.Mod: lambda a, b: a % b,
    ast.Pow: _safe_pow,
}

_UNARY = {ast.UAdd: lambda a: +a, ast.USub: lambda a: -a}


def _eval_node(node, table):
    """`table` = (headers, rows) en mode fichier, sinon None."""
    if isinstance(node, ast.Expression):
        return _eval_node(node.body, table)
    if isinstance(node, ast.Constant):
        if isinstance(node.value, (int, float)) and not isinstance(node.value, bool):
            return node.value
        if isinstance(node.value, str):
            if table is None:
                raise ToolError(
                    f"chaîne {node.value!r} : les références de colonnes exigent le "
                    "paramètre `file` (sinon l'expression doit être purement numérique)"
                )
            return _column(table[0], table[1], node.value)
        raise ToolError(f"constante non supportée : {node.value!r}")
    if isinstance(node, ast.BinOp) and type(node.op) in _BINOPS:
        left = _eval_node(node.left, table)
        right = _eval_node(node.right, table)
        _reject_col(left)
        _reject_col(right)
        return _BINOPS[type(node.op)](left, right)
    if isinstance(node, ast.UnaryOp) and type(node.op) in _UNARY:
        val = _eval_node(node.operand, table)
        _reject_col(val)
        return _UNARY[type(node.op)](val)
    if isinstance(node, ast.Name):
        if node.id in _CONSTS:
            return _CONSTS[node.id]
        raise ToolError(
            f"nom inconnu : {node.id} (constantes admises : {', '.join(_CONSTS)})"
        )
    if isinstance(node, ast.Call):
        if not isinstance(node.func, ast.Name) or node.keywords:
            raise ToolError("appel non supporté")
        fname = node.func.id
        args = [_eval_node(a, table) for a in node.args]
        # Agrégat sur colonne : sum("Débit"), count("Compte")…
        if len(args) == 1 and isinstance(args[0], _Col):
            if fname not in _AGGS:
                raise ToolError(
                    f"{fname} n'accepte pas une colonne (agrégats : "
                    f"{', '.join(sorted(_AGGS))})"
                )
            col = args[0]
            if fname == "count":
                return float(
                    sum(1 for c in col.values if c not in (None, "") and str(c).strip())
                )
            nums = [n for n in (_to_number(c) for c in col.values) if n is not None]
            if not nums and fname in ("min", "max"):
                raise ToolError(f"colonne {col.name!r} : aucune valeur numérique")
            return float(_AGGS[fname](nums)) if nums else 0.0
        for a in args:
            _reject_col(a)
        if fname in _FUNCS:
            return _FUNCS[fname](*args)
        raise ToolError(
            f"fonction non autorisée : {fname} (admises : {', '.join(sorted(_FUNCS))})"
        )
    # Cas fréquent : une virgule décimale FR (« 1,5 ») ou un séparateur de milliers
    # (« 1,000 ») fait lire l'expression comme un tuple par Python. Message actionnable
    # plutôt que « syntaxe non supportée : Tuple » (la virgule ne sépare QUE les
    # arguments de fonction ici).
    if isinstance(node, ast.Tuple):
        raise ToolError(
            "virgule inattendue : utilise un POINT décimal (1.5, pas 1,5) et pas de "
            "séparateur de milliers (1000, pas 1,000). La virgule ne sépare que les "
            "arguments d'une fonction, ex. log(100, 10)."
        )
    raise ToolError(f"syntaxe non supportée : {type(node).__name__}")


def _reject_col(v) -> None:
    if isinstance(v, _Col):
        raise ToolError(
            f"la colonne {v.name!r} doit passer par un agrégat "
            f"({', '.join(sorted(_AGGS))}) avant d'entrer dans un calcul"
        )


def _fmt(value) -> str:
    """Résultat exact d'abord ; arrondi lisible en plus quand le float est verbeux."""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = repr(value)
    if isinstance(value, float) and len(text) > 12:
        # ASCII uniquement (consoles Windows cp1252 + préférence style du projet).
        return f"{text} (~ {value:,.4f})".replace(",", " ")
    return text


def calculate(
    expression: str,
    file: str = "",
    where: dict | None = None,
    workspace_dir: str = ".",
) -> str:
    expression = (expression or "").strip()
    if not expression:
        raise ToolError("argument 'expression' manquant")
    if len(expression) > 2000:
        raise ToolError("expression trop longue (2000 caractères max)")
    table = None
    info = ""
    if file:
        path = _resolve_in_root(Path(workspace_dir), file)
        headers, rows = _load_table(path)
        if where:
            rows = _apply_where(headers, rows, where)
        table = (headers, rows)
        info = f"  [{Path(file).name} : {len(rows)} ligne(s) considérée(s)]"
    # Traduit ^, exposants Unicode, ×÷− etc. en syntaxe Python (l'affichage garde
    # l'écriture ORIGINALE du modèle, plus lisible et sans ** qui grasseraient en markdown).
    normalized = _normalize_expr(expression)
    try:
        tree = ast.parse(normalized, mode="eval")
    except SyntaxError as exc:
        raise ToolError(f"expression invalide : {exc.msg}") from exc
    try:
        result = _eval_node(tree, table)
    except ToolError:
        raise
    except ZeroDivisionError as exc:
        raise ToolError("division par zéro") from exc
    except (OverflowError, ValueError) as exc:
        raise ToolError(f"calcul impossible : {exc}") from exc
    _reject_col(result)
    return f"{expression} = {_fmt(result)}{info}"


def make_calculate(workspace_dir: str = ".") -> ToolSpec:
    return ToolSpec(
        name="calculate",
        description=(
            "Evaluates arithmetic EXACTLY. ALWAYS use it for any calculation beyond a "
            "single trivial operation — money, percentages, VAT, totals — and NEVER "
            "compute multi-step arithmetic in your head: language models routinely "
            "get sums wrong. Plain expressions: (55000-600)*0.2, 2^10, sqrt/log/trig "
            "and usual math functions accepted. With `file` (CSV/TSV/XLSX), quoted "
            'strings become COLUMN references with aggregates — sum("Debit") - '
            'sum("Credit"), avg, count — and `where` filters rows first: NEVER sum '
            "table rows mentally."
        ),
        parameters={
            "type": "object",
            "properties": {
                "expression": {
                    "type": "string",
                    "description": (
                        "Arithmetic expression, e.g. (55000-600)*0.2 — or with file: "
                        'sum("Debit") - sum("Credit")'
                    ),
                },
                "file": {
                    "type": "string",
                    "description": (
                        "Optional CSV/TSV/XLSX path (workspace-relative). Enables "
                        "column references and aggregates."
                    ),
                },
                "where": {
                    "type": "object",
                    "description": (
                        'Optional row filter before aggregating: {"column": "Compte", '
                        '"equals": "512"}'
                    ),
                    "properties": {
                        "column": {"type": "string"},
                        "equals": {},
                    },
                },
            },
            "required": ["expression"],
        },
        run=lambda args: calculate(
            args.get("expression", ""),
            file=args.get("file", "") or "",
            where=args.get("where") or None,
            workspace_dir=workspace_dir,
        ),
    )
